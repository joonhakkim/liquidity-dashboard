"""
data/manual/*데이터 모음*.xlsm (에프앤가이드류 컨센서스 워크북)에서
- '시가총액' 시트: 종목별 시가총액
- '2026 영업이익 추정치' / '2027 영업이익 추정치' 시트: 종목별 영업이익 컨센서스(각 시트 최신일자 행)
을 읽어 시가총액이 있는 전체 상장사를 data/screening/op_growth_screen.csv 로 저장한다.
컨센서스(영업이익 추정치)는 애널리스트 커버리지가 있는 종목(~800개)만 있어서 없는 종목은
2026/2027 영업이익·증가율 컬럼이 비어있다 - 컨센서스가 없어도 DART 실적 트래킹은 하기 위해
시가총액만 있으면 포함시킨다(이름 기준 아우터조인).

주의: 이 워크북엔 매출액 컨센서스가 없어 영업이익만으로 스크리닝한다(사용자 확인 완료).

'시가총액' 시트 자체가 데이터터미널에서 필터링된 상태로 내보내져서 삼성전자를 포함한
800여개 종목이 통째로 빠져있는 경우가 있었다(2026-08-16, 사용자가 "삼성전자가 없어졌다"고
알려줘서 발견 - 예전에 수동으로 "이전 스냅샷 값 보완" 패치를 한 적 있는데 스크립트 자체엔
반영이 안 돼서 매일 재실행될 때마다 다시 빠지는 문제가 있었음). 이번엔 코드 자체에
KRX 공식 Open API(fetch_adr.py와 동일 소스)로 전종목 시가총액을 받아와서, 워크북에
없는 종목만 이걸로 보완하도록 고쳤다 - 워크북 상태와 무관하게 매일 자동으로 채워진다.
"""
import glob
import os
import re
import time
import warnings
from datetime import datetime, timedelta

import openpyxl
import pandas as pd
import requests
from dotenv import load_dotenv

warnings.filterwarnings("ignore")

load_dotenv()
KRX_OPEN_API_KEY = os.environ.get("KRX_OPEN_API_KEY")
KRX_STK_URL = "https://data-dbg.krx.co.kr/svc/apis/sto/stk_bydd_trd"
KRX_KSQ_URL = "https://data-dbg.krx.co.kr/svc/apis/sto/ksq_bydd_trd"

EXCLUDE_NAME_RE = re.compile(r"스팩|기업인수목적|ETF|ETN$")

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
MANUAL_DIR = os.path.join(DATA_DIR, "manual")
OUT_PATH = os.path.join(DATA_DIR, "screening", "op_growth_screen.csv")

MIN_MARKET_CAP = 0  # 시총 제한 없음 - 컨센서스(영업이익 추정치) 있는 종목 전체(~740개)가 사실상 상한
MIN_OP_GROWTH = -10  # 사실상 무제한(적자전환 등 극단치만 배제) - 실제 필터링은 웹페이지에서
# DART 일일 API 호출 한도(2만 건) 때문에 하루에 전체를 다 못 돈다 - fetch_dart_quarterly.py /
# fetch_valuation_bands.py / fetch_dart_preliminary.py 가 이미 처리된 종목은 건너뛰므로,
# 하루 2회 자동 실행(run_pipeline.py)이 며칠 반복되면 전체가 자연스럽게 채워진다.


def find_workbook():
    candidates = glob.glob(os.path.join(MANUAL_DIR, "*데이터 모음*.xls*"))
    candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]
    if not candidates:
        raise SystemExit("data/manual/ 에 '데이터 모음' 워크북이 없습니다.")
    return max(candidates, key=os.path.getmtime)


def read_market_cap(wb):
    ws = wb["시가총액"]
    rows = list(ws.iter_rows(values_only=True))
    code_row_idx = next(
        i for i, r in enumerate(rows) if r[0] == "Code" and isinstance(r[1], str) and r[1].startswith("A")
    )
    names = rows[code_row_idx + 1]
    date_row_idx = next(i for i in range(code_row_idx + 2, len(rows)) if rows[i][0] == "D A T E") + 1
    latest = rows[date_row_idx]
    print(f"  시가총액 기준일: {latest[0]}")
    return {names[i]: latest[i] for i in range(1, len(names)) if names[i] and latest[i] is not None}


def fetch_krx_market_cap():
    """KRX 공식 Open API(fetch_adr.py와 동일 소스)로 코스피+코스닥 전종목 시가총액을 받아온다.
    최근 영업일부터 거꾸로 최대 10일 재시도(휴장일 등으로 데이터 없는 날 건너뛰기)."""
    if not KRX_OPEN_API_KEY:
        print("  경고: KRX_OPEN_API_KEY가 없어 시가총액 보완을 건너뜁니다.")
        return {}

    headers = {"AUTH_KEY": KRX_OPEN_API_KEY}
    d = datetime.today()
    result = {}
    for _ in range(10):
        bas_dd = d.strftime("%Y%m%d")
        got_any = False
        for url in (KRX_STK_URL, KRX_KSQ_URL):
            try:
                r = requests.get(url, params={"basDd": bas_dd}, headers=headers, timeout=20)
                data = r.json().get("OutBlock_1", []) if r.status_code == 200 else []
            except requests.RequestException:
                data = []
            if data:
                got_any = True
            for item in data:
                name = item.get("ISU_NM")
                mktcap = pd.to_numeric(item.get("MKTCAP"), errors="coerce")
                if name and pd.notna(mktcap):
                    result[name] = float(mktcap)  # numpy.int64는 isinstance(v, (int,float))를 안 통과해서 순수 float로
            time.sleep(0.1)
        if got_any:
            print(f"  KRX API 기준일: {bas_dd} ({len(result)}종목)")
            break
        d -= timedelta(days=1)
    return result


def read_op_estimate(wb, sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    names = rows[1]
    latest = rows[2]  # 날짜 내림차순 정렬이라 2행이 최신
    print(f"  {sheet_name} 기준일: {latest[0]}")
    return {
        names[i]: latest[i]
        for i in range(1, len(names))
        if names[i] and names[i] != "합계" and latest[i] is not None
    }


def main():
    path = find_workbook()
    print(f"워크북 로딩: {os.path.basename(path)}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)

    print("시가총액 읽는 중...")
    market_cap = read_market_cap(wb)
    print(f"  {len(market_cap)}개 종목")

    print("KRX 공식 API로 시가총액 보완 중(워크북에 누락된 종목만 채움)...")
    krx_market_cap = fetch_krx_market_cap()
    added = 0
    for name, cap in krx_market_cap.items():
        if name not in market_cap:
            market_cap[name] = cap
            added += 1
    print(f"  {added}개 종목 보완됨 (워크북 {len(market_cap) - added}개 + KRX 보완 {added}개 = 총 {len(market_cap)}개)")

    print("2026 영업이익 추정치 읽는 중...")
    op_2026 = read_op_estimate(wb, "2026 영업이익 추정치")
    print("2027 영업이익 추정치 읽는 중...")
    op_2027 = read_op_estimate(wb, "2027 영업이익 추정치")

    # 컨센서스(영업이익 추정치)가 없어도 시가총액만 있으면 포함한다 - 성장률/컨센서스 필터는
    # 페이지에서 걸고, 여기서는 "실적 트래킹 대상"을 최대한 넓게(전체 상장사) 잡는다.
    # 컨센서스가 있는 종목만 OP2026/OP2027/증가율이 채워지고, 없는 종목은 None으로 둔다.
    rows = []
    for name, cap in market_cap.items():
        if not isinstance(cap, (int, float)):
            continue
        if EXCLUDE_NAME_RE.search(name):
            continue  # 스팩(SPAC)·ETF·ETN 등 실적 추적 대상이 아닌 종목 제외
        v2026 = op_2026.get(name)
        v2027 = op_2027.get(name)
        growth = None
        if isinstance(v2026, (int, float)) and isinstance(v2027, (int, float)) and v2026 > 0:
            growth = round((v2027 - v2026) / v2026, 4)
        rows.append({
            "종목명": name,
            "시가총액(억원)": round(cap / 1e8, 1),
            "2026_영업이익(십억원)": round(v2026, 1) if isinstance(v2026, (int, float)) else None,
            "2027_영업이익(십억원)": round(v2027, 1) if isinstance(v2027, (int, float)) else None,
            "영업이익_증가율": growth,
        })

    df = pd.DataFrame(rows)
    screened = df[df["시가총액(억원)"] >= MIN_MARKET_CAP / 1e8]
    # 시가총액 큰 종목부터 먼저 처리되도록 정렬한다 - DART 백필이 며칠 걸리는데(2500여개),
    # 성장률 순으로 두면 삼성전자 같은 대형주가 한참 뒤로 밀려서 오래 기다려야 했다.
    # 페이지 자체의 기본 정렬(OP 증가율)은 화면(JS)에서 따로 적용되므로 순서를 바꿔도 무방.
    screened = screened.sort_values(
        "시가총액(억원)", ascending=False, na_position="last"
    ).reset_index(drop=True)

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    screened.to_csv(OUT_PATH, index=False, encoding="utf-8-sig")

    has_consensus = screened["영업이익_증가율"].notna().sum()
    print(f"\n전체 대상(시가총액 있는 전 종목): {len(screened)}개")
    print(f"이 중 영업이익 컨센서스(2026/2027 둘 다 흑자) 있는 종목: {has_consensus}개")
    print(f"저장 완료: {OUT_PATH}")


if __name__ == "__main__":
    main()
