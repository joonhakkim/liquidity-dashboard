"""
OP밴드(영업이익 밴드) 트래커 - data/manual/*기업*밴드*.xlsx(데이터터미널 시가총액/영업이익
컨센서스 export)를 읽어서 종목별 "영업이익 추정치 x N배" 밴드 차트 데이터를 만든다.

워크북 구조(사용자 제공 스펙) - 신형/구형 두 포맷을 자동 감지해서 둘 다 처리한다:
- 구형("기업 밴드 찾기.xlsx", 시트 "구성 1/2/3" - 엑셀 컬럼 한계 16,384열 때문에 종목을 나눠
  이어붙인 것): 종목 1개당 +0 시가총액(S102100) / +1 TTM(M121505.M) / +2~ 분기별 영업이익
  추정치(E121500.M, 4개씩 묶여 회계연도 1개, Base Date=YYYYMM)
- 신형("기업 밴드 찾기 2.xlsx"부터, 시트 "기업밴드" 단일 - 2026-09-03 도입, 시계열이
  2021-12-30~로 더 길고 분기합산 없이 FnGuide가 이미 집계한 연간 직접추정치를 줌 - 분기합
  대비 오차 ±1~3% 수준(2026-09-03 확인, 삼성전자 등 대형주도 이상치 아님으로 판단하고 그대로
  사용): 종목 1개당 +0 시가총액(S102100) / +1 TTM(M121505.M) / +2 당해연도 추정(E121500.M,
  Base Date="NFY1") / +3 차년도 추정(E121500.M, Base Date="NFY2")
- 공통: 8행=종목코드, 9행=종목명, 10행=item code, 11행=단위, 12행=Base Date, 15행부터
  날짜별 데이터(A열=날짜)

산출 로직:
1) 코드가 연속되는 컬럼 구간을 자동 탐지해서 회사 블록으로 인식(블록 폭 하드코딩 안 함)
2) item code + Base Date로 역할 컬럼(시총/TTM/분기 or NFY1·NFY2) 탐지 - 구형이면 분기 그룹핑,
   신형이면 NFY1/NFY2 컬럼을 직접 사용
3) 날짜별로 "6월 스위칭 룰"로 쓸 회계연도 결정(1~6월->해당연도, 7~12월->다음연도) - 구형은
   해당 FY의 분기 4개가 전부 있어야 합산, 신형은 use_year가 그 날짜의 해(NFY1) 또는 다음 해
   (NFY2)인지에 따라 해당 컬럼을 그대로 씀(신형 NFY1=그 날짜가 속한 연도, NFY2=다음 연도로
   확인됨). 어느 포맷이든 "6월 기준으로 어느 연도를 쓸지"의 최종 결과는 동일하다(사용자 확인).
4) 위에서 못 구하면 TTM 컬럼값(1순위) -> 직전에 계산된 값 이어쓰기(2순위) -> 그래도 없으면 스킵
5) 배수 = 시가총액 / (영업이익추정치 x 1000)  (영업이익은 천원 단위라 원 단위로 환산)
6) "보기 좋은" 밴드 배수 자동 선정(1,2,5,10,15,20,25,30,50,100,200,250,500,1000 중에서
   목표 라인수에 맞는 간격을 골라 균등 배치) - 밴드 선 자체는 화면(JS)에서 op(t) x 배수로 그린다
   (원본 시계열을 밴드 개수만큼 중복 저장할 필요가 없어서 종목별 JSON이 훨씬 가벼워진다)

출력:
- docs/op_band_data/<종목코드>.json: 종목별 시계열(날짜/시총/영업이익/배수) + 밴드 배수 목록(차트용)
- data/screening/op_band_summary.csv + docs/op_band.html 안에 임베드되는 요약 테이블(필터/정렬용):
  최신 배수, 과거 밴드(10~90퍼센타일) 대비 지금이 몇 퍼센타일인지 - 지금 밴드 하단에 가까운(=
  이익 대비 시총이 눌린) 종목을 쉽게 찾을 수 있게. 단, 기간중 최소/최대/백분위는 흑자(양수 배수)
  구간만으로 계산한다(적자 구간은 배수가 음수로 튀어서 최소치가 의미 없는 숫자가 됨).
"""
import glob
import json
import math
import os
from datetime import datetime

import openpyxl
import pandas as pd

from build_screening_page import load_sector_map

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
MANUAL_DIR = os.path.join(DATA_DIR, "manual")
SCREEN_DIR = os.path.join(DATA_DIR, "screening")
DOCS_DIR = os.path.join(os.path.dirname(__file__), "..", "docs")
DETAIL_OUT_DIR = os.path.join(DOCS_DIR, "op_band_data")
SUMMARY_PATH = os.path.join(SCREEN_DIR, "op_band_summary.csv")
PAGE_OUT_PATH = os.path.join(DOCS_DIR, "op_band.html")
FNGUIDE_PATH = os.path.join(DATA_DIR, "op_band_fnguide.csv")

MKTCAP_ITEM = "S102100"
TTM_ITEM = "M121505.M"
QUARTER_ITEM = "E121500.M"
HEADER_CODE_ROW = 8
HEADER_NAME_ROW = 9
HEADER_ITEM_ROW = 10
HEADER_BASEDATE_ROW = 12
DATA_START_ROW = 15

NICE_STEPS = [1, 2, 5, 10, 15, 20, 25, 30, 50, 100, 200, 250, 500, 1000]
TARGET_BAND_LINES = 6


def find_workbook():
    candidates = glob.glob(os.path.join(MANUAL_DIR, "*기업*밴드*.xls*"))
    candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)


def detect_blocks(row_codes):
    """1번(A열=날짜/라벨열) 다음부터 코드가 연속되는 구간을 회사 블록으로 묶는다."""
    blocks = []
    n = len(row_codes)
    i = 1
    while i < n:
        code = row_codes[i]
        if code is None:
            i += 1
            continue
        j = i
        while j < n and row_codes[j] == code:
            j += 1
        blocks.append((code, i, j))
        i = j
    return blocks


def process_sheet(ws):
    max_col = ws.max_column
    max_row = ws.max_row

    def read_row(r):
        return next(ws.iter_rows(min_row=r, max_row=r, max_col=max_col, values_only=True))

    row_codes = read_row(HEADER_CODE_ROW)
    row_names = read_row(HEADER_NAME_ROW)
    row_items = read_row(HEADER_ITEM_ROW)
    row_basedate = read_row(HEADER_BASEDATE_ROW)

    blocks = detect_blocks(row_codes)

    dates = []
    data_rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, max_row=max_row, max_col=max_col, values_only=True):
        d = row[0]
        if d is None:
            continue
        if isinstance(d, datetime):
            dates.append(d)
            data_rows.append(row)

    results = {}
    for code, start, end in blocks:
        mktcap_idx = ttm_idx = nfy1_idx = nfy2_idx = None
        quarter_idxs = []
        for k in range(start, end):
            item = row_items[k]
            base = row_basedate[k]
            if item == MKTCAP_ITEM:
                mktcap_idx = k
            elif item == TTM_ITEM:
                ttm_idx = k
            elif item == QUARTER_ITEM:
                # "기업 밴드 찾기 2" 신형식은 이 item code 컬럼이 분기(YYYYMM) 대신 NFY1/NFY2
                # 라벨의 "연간 직접추정치" 컬럼 2개뿐이다(2026-09-03, 시계열이 더 길고 분기
                # 합산 없이 바로 연간 추정치를 준다 - 사용자 확인: 분기합과 오차 ±1~3%
                # 수준이라 대체 가능). 라벨로 신/구형식을 구분해서 같은 함수가 둘 다 처리한다.
                if base == "NFY1":
                    nfy1_idx = k
                elif base == "NFY2":
                    nfy2_idx = k
                elif base is not None:
                    quarter_idxs.append(k)

        if mktcap_idx is None or (not quarter_idxs and nfy1_idx is None and nfy2_idx is None):
            continue

        name = row_names[start] if start < len(row_names) else code

        fy_blocks = {}
        n_full_groups = len(quarter_idxs) - (len(quarter_idxs) % 4)
        for gi in range(0, n_full_groups, 4):
            group = quarter_idxs[gi:gi + 4]
            base = row_basedate[group[0]]
            if base is None:
                continue
            try:
                year = int(base) // 100
            except (TypeError, ValueError):
                continue
            fy_blocks[year] = group

        series_dates, series_mktcap, series_op, series_mult = [], [], [], []
        prev_op = None
        for row_vals, d in zip(data_rows, dates):
            mktcap = row_vals[mktcap_idx]
            if mktcap is None:
                continue
            # 6월 스위칭 룰: 1~6월엔 그 해, 7~12월엔 다음 해 영업이익 추정치를 쓴다(신구 형식
            # 공통 - 회계연도 표시 방식만 다를 뿐 어느 연도를 쓸지의 기준 자체는 동일해야 함,
            # 2026-09-03 사용자 확인). 신형식은 NFY1=그 날짜가 속한 연도, NFY2=다음 연도라
            # (2025-03/2026-08 시점 값으로 검증) use_year==d.year일 때 NFY1, use_year==d.year+1일
            # 때 NFY2를 그대로 쓰면 6월 스위칭과 동일한 결과가 된다 - 별도 분기합산 불필요.
            use_year = d.year if d.month <= 6 else d.year + 1
            op = None
            if fy_blocks:
                group = fy_blocks.get(use_year)
                if group:
                    vals = [row_vals[gi] for gi in group]
                    if all(v is not None for v in vals):
                        op = sum(vals)
            elif use_year == d.year and nfy1_idx is not None:
                op = row_vals[nfy1_idx]
            elif use_year == d.year + 1 and nfy2_idx is not None:
                op = row_vals[nfy2_idx]
            if op is None and ttm_idx is not None:
                op = row_vals[ttm_idx]
            if op is None:
                op = prev_op
            if op is None or op == 0:
                continue
            prev_op = op

            op_won = op * 1000  # 천원 -> 원
            mult = mktcap / op_won
            series_dates.append(d.strftime("%Y-%m-%d"))
            series_mktcap.append(round(mktcap, 0))
            series_op.append(round(op_won, 0))
            series_mult.append(round(mult, 4))

        if not series_dates:
            continue

        results[code] = {"name": name, "dates": series_dates, "mktcap": series_mktcap,
                          "op": series_op, "mult": series_mult}
    return results


def pick_band_multiples(mults):
    positive = [m for m in mults if m and m > 0]
    if not positive:
        return []
    end = math.ceil(max(positive)) + 1
    threshold = end / TARGET_BAND_LINES
    step = next((s for s in NICE_STEPS if s >= threshold), NICE_STEPS[-1])
    multiples = []
    v = step
    while v <= end:
        multiples.append(v)
        v += step
    return multiples or [step]


def load_naver_sector_map():
    """data/sector_map.csv(fetch_naver_sector.py 결과) -> {종목코드(6자리): 섹터}.
    네이버 자체 업종분류라 상장사 대부분(2500여종목 중 2519개, 99%)을 커버한다 - 반면
    load_sector_map()(엑셀 "섹터별 구성 종목" 시트)는 팀이 관리하는 테마성 큐레이션이라 913개
    종목만 커버. 그래서 이쪽을 1순위로 쓰고, 여기 없는 종목만 엑셀 매핑으로 보충한다."""
    path = os.path.join(DATA_DIR, "sector_map.csv")
    if not os.path.exists(path):
        return {}
    df = pd.read_csv(path, dtype={"code": str})
    return dict(zip(df["code"], df["sector"]))


def build_summary_row(code, data, sector_map, naver_sector_map):
    """기간중 최소/최대/백분위는 흑자(영업이익 양수, mult>0) 구간만으로 계산한다 - 적자 구간은
    배수가 음수로 튀어서(예: 시총은 그대로인데 영업이익이 살짝만 적자여도 배수가 -수백~수천배로
    뜀) "기간중 최소"가 실제 밸류에이션 하단과 무관한 의미 없는 숫자가 되는 문제가 있었다
    (2026-08-26, 사용자가 "괴리가 너무 크다"고 지적). 현재(latest_mult)는 적자여도 있는 그대로
    보여준다 - 실제로 지금 적자인 건 사실이니까. 흑자였던 적이 아예 없는 종목은 밴드 자체가
    의미 없으니 행을 생성하지 않는다(값을 지어내지 않는다)."""
    mults = [m for m in data["mult"] if m is not None]
    if not mults:
        return None
    latest_mult = mults[-1]
    positive_mults = [m for m in mults if m > 0]
    if not positive_mults:
        return None
    sorted_mults = sorted(positive_mults)
    below = sum(1 for m in sorted_mults if m <= latest_mult)
    percentile = below / len(sorted_mults) * 100
    bare_code = code.lstrip("A")
    sector = naver_sector_map.get(bare_code) or sector_map.get(data["name"])
    return {
        "code": code, "name": data["name"],
        "sector": sector,
        "latest_date": data["dates"][-1],
        "latest_mult": round(latest_mult, 2),
        "hist_min_mult": round(sorted_mults[0], 2),
        "hist_max_mult": round(sorted_mults[-1], 2),
        "percentile": round(percentile, 1),
        "n_obs": len(positive_mults),
        "latest_mktcap": round(data["mktcap"][-1], 0) if data["mktcap"] else None,
    }


def load_fnguide_map():
    """data/op_band_fnguide.csv(fetch_op_band_consensus.py 결과) -> {code: {year: op_won}}."""
    if not os.path.exists(FNGUIDE_PATH):
        return {}
    df = pd.read_csv(FNGUIDE_PATH, dtype={"code": str})
    m = {}
    for _, row in df.iterrows():
        m.setdefault(row["code"], {})[int(row["year"])] = row["op_100mil"] * 1e8  # 억원 -> 원
    return m


def load_manual_overrides():
    """data/manual/op_band_overrides.csv -> {code: {year: op_won}} - 사용자가 직접 지정한 값.
    FnGuide보다도 우선순위가 높다(사용자가 명시적으로 지정한 값이 최종 권위)."""
    path = os.path.join(MANUAL_DIR, "op_band_overrides.csv")
    if not os.path.exists(path):
        return {}
    df = pd.read_csv(path, dtype={"code": str})
    m = {}
    for _, row in df.iterrows():
        m.setdefault(row["code"], {})[int(row["year"])] = row["op_100mil"] * 1e8  # 억원 -> 원
    return m


def apply_year_override(code, data, override_map, current_use_year):
    """엑셀/FnGuide 값을 현재 회계연도(use_year) 구간에 한해 override_map 값으로 덮어쓴다.
    과거 구간(이미 지나간 회계연도)은 되돌릴 수 없으니 건드리지 않고, 지금 활성 회계연도에
    해당하는 날짜들(시계열 맨 끝의 연속 구간)만 보정한다. FnGuide 교차검증과 사용자 수동
    지정(data/manual/op_band_overrides.csv) 둘 다 이 함수를 공용으로 쓴다."""
    op_by_year = override_map.get(code)
    if not op_by_year or current_use_year not in op_by_year:
        return data

    new_op_won = op_by_year[current_use_year]
    if new_op_won == 0:
        return data

    dates = data["dates"]
    mktcap = data["mktcap"]
    op = data["op"][:]
    mult = data["mult"][:]

    for i in range(len(dates) - 1, -1, -1):
        d = datetime.strptime(dates[i], "%Y-%m-%d")
        use_year = d.year if d.month <= 6 else d.year + 1
        if use_year != current_use_year:
            break
        op[i] = round(new_op_won, 0)
        if mktcap[i] is not None:
            mult[i] = round(mktcap[i] / new_op_won, 4)

    data["op"] = op
    data["mult"] = mult
    return data


def main():
    wb_path = find_workbook()
    if not wb_path:
        print("data/manual/ 에 '기업 밴드' 워크북이 없습니다.")
        return

    print(f"워크북 로드 중: {wb_path}")
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)

    # 워크북의 시트를 전부 순회 - 구형(엑셀 컬럼 한계 때문에 "구성 1/2/3"로 나뉜 파일)/
    # 신형("기업밴드" 단일 시트, 2026-09-03 "기업 밴드 찾기 2"부터) 둘 다 이름 하드코딩 없이
    # 자동으로 처리된다.
    all_results = {}
    for sn in wb.sheetnames:
        print(f"처리 중: {sn} ...")
        results = process_sheet(wb[sn])
        for code, data in results.items():
            if code not in all_results:
                all_results[code] = data
        print(f"  {len(results)}개 종목 처리")

    print(f"\n총 {len(all_results)}개 종목")

    fnguide_map = load_fnguide_map()
    today = datetime.today()
    current_use_year = today.year if today.month <= 6 else today.year + 1
    if fnguide_map:
        print(f"FnGuide 보정 적용 중(현재 회계연도 FY{current_use_year}, {len(fnguide_map)}종목 커버)...")
        for code in all_results:
            all_results[code] = apply_year_override(code, all_results[code], fnguide_map, current_use_year)
    else:
        print("FnGuide 데이터 없음(fetch_op_band_consensus.py 미실행) - 엑셀 원본만 사용")

    manual_map = load_manual_overrides()
    if manual_map:
        print(f"수동 지정값 적용 중({len(manual_map)}종목, FnGuide보다 우선)...")
        for code in manual_map:
            if code in all_results:
                all_results[code] = apply_year_override(code, all_results[code], manual_map, current_use_year)

    os.makedirs(DETAIL_OUT_DIR, exist_ok=True)
    os.makedirs(SCREEN_DIR, exist_ok=True)

    # 섹터: 1순위는 네이버 업종분류(fetch_naver_sector.py, 종목코드 기준, 상장사 대부분 커버),
    # 2순위는 주식 스크리닝 페이지가 쓰는 "*데이터 모음*.xlsm"의 "섹터별 구성 종목" 시트(팀이
    # 관리하는 테마성 큐레이션, 종목명 기준, 913종목만 커버) - 네이버 쪽에 없는 종목만 보충.
    naver_sector_map = load_naver_sector_map()
    sector_map = load_sector_map()
    print(f"네이버 업종 매핑 {len(naver_sector_map)}종목, 큐레이션 섹터 매핑 {len(sector_map)}종목 로드됨")

    summary_rows = []
    for code, data in all_results.items():
        band_multiples = pick_band_multiples(data["mult"])
        detail = {
            "code": code, "name": data["name"],
            "dates": data["dates"], "mktcap": data["mktcap"],
            "op": data["op"], "mult": data["mult"],
            "bandMultiples": band_multiples,
        }
        with open(os.path.join(DETAIL_OUT_DIR, f"{code}.json"), "w", encoding="utf-8") as f:
            json.dump(detail, f, ensure_ascii=False)

        row = build_summary_row(code, data, sector_map, naver_sector_map)
        if row:
            summary_rows.append(row)

    summary_df = pd.DataFrame(summary_rows).sort_values("latest_mult")
    summary_df.to_csv(SUMMARY_PATH, index=False, encoding="utf-8-sig")
    print(f"저장 완료: {SUMMARY_PATH} ({len(summary_df)}종목)")

    build_page(summary_rows, wb_path)


def build_page(summary_rows, wb_path):
    updated_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    src_mtime = datetime.fromtimestamp(os.path.getmtime(wb_path)).strftime("%Y-%m-%d %H:%M")
    html = TEMPLATE.format(
        rows_json=json.dumps(summary_rows, ensure_ascii=False),
        updated_at=updated_at,
        src_mtime=src_mtime,
        n_stocks=len(summary_rows),
    )
    with open(PAGE_OUT_PATH, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"저장 완료: {PAGE_OUT_PATH}")


TEMPLATE = """<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>OP밴드 트래커</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
  body {{ font-family: -apple-system, "Malgun Gothic", sans-serif; background:#0f1115; color:#e6e6e6; margin:0; padding:24px; }}
  a.back {{ color:#4dabf7; font-size:13px; text-decoration:none; margin-right:12px; }}
  h1 {{ font-size:20px; margin:8px 0 4px 0; }}
  .updated {{ color:#9aa0a6; font-size:13px; margin-bottom:20px; }}
  .filters {{ display:flex; gap:12px; flex-wrap:wrap; align-items:center; margin-bottom:16px; font-size:13px; color:#9aa0a6; }}
  .filters input, .filters select {{ background:#1a1d24; border:1px solid #2a2e37; color:#e6e6e6; border-radius:6px; padding:6px 10px; font-size:13px; font-family:inherit; }}
  .filters input[type="text"] {{ width:160px; }}
  .filters input[type="number"] {{ width:80px; }}
  .filters label {{ display:flex; align-items:center; gap:6px; }}
  table {{ border-collapse: collapse; width:100%; max-width:1000px; font-size:13px; }}
  th, td {{ padding:8px 12px; text-align:right; border-bottom:1px solid #23262e; }}
  th:first-child, td:first-child {{ text-align:left; }}
  th:nth-child(2), td:nth-child(2), th:nth-child(3), td:nth-child(3) {{ text-align:left; color:#9aa0a6; }}
  th {{ color:#9aa0a6; font-weight:normal; font-size:12px; cursor:pointer; user-select:none; }}
  th:hover {{ color:#c7cbd1; }}
  tr {{ cursor:pointer; }}
  tr:hover {{ background:#1a1d24; }}
  .pctl-low {{ color:#4dabf7; font-weight:bold; }}
  .pctl-high {{ color:#ff6b6b; }}
  .overlay {{ display:none; position:fixed; inset:0; background:rgba(0,0,0,0.6); z-index:10; align-items:flex-start; justify-content:center; padding:40px 20px; overflow-y:auto; }}
  .overlay.open {{ display:flex; }}
  .detail-card {{ background:#14161c; border:1px solid #23262e; border-radius:14px; padding:24px; max-width:900px; width:100%; }}
  .detail-card h2 {{ margin:0 0 4px 0; font-size:18px; }}
  .detail-card .sub {{ color:#9aa0a6; font-size:13px; margin-bottom:16px; }}
  .range-bar {{ display:flex; gap:6px; margin-bottom:10px; flex-wrap:wrap; }}
  .range-btn {{ background:#1a1d24; border:1px solid #2a2e37; color:#9aa0a6; padding:5px 12px;
    border-radius:999px; cursor:pointer; font-size:12px; font-family:inherit; }}
  .range-btn:hover {{ color:#c7cbd1; border-color:#4dabf7; }}
  .range-btn.active {{ background:#4dabf7; color:#0f1115; border-color:#4dabf7; font-weight:bold; }}
  .custom-range-bar {{ display:flex; align-items:center; gap:10px; margin-bottom:14px; flex-wrap:wrap; font-size:13px; color:#9aa0a6; }}
  .custom-range-bar label {{ display:flex; align-items:center; gap:6px; }}
  .custom-range-bar input[type="date"] {{
    background:#1a1d24; border:1px solid #2a2e37; color:#e6e6e6; border-radius:6px; padding:5px 8px; font-size:13px;
  }}
  .custom-range-bar button {{ background:#1a1d24; border:1px solid #2a2e37; color:#9aa0a6; padding:6px 14px;
    border-radius:6px; cursor:pointer; font-size:12px; font-family:inherit; }}
  .custom-range-hint {{ color:#63e6be; }}
  .band-controls {{ display:flex; gap:12px; flex-wrap:wrap; align-items:center; margin-bottom:12px; font-size:13px; color:#9aa0a6; }}
  .band-controls label {{ display:flex; align-items:center; gap:6px; }}
  .band-controls input {{ background:#1a1d24; border:1px solid #2a2e37; color:#e6e6e6; border-radius:6px; padding:5px 8px; font-size:13px; font-family:inherit; }}
  .band-controls input[type="number"] {{ width:60px; }}
  .band-controls input[type="text"] {{ width:180px; }}
  .band-controls button {{ background:#1a1d24; border:1px solid #2a2e37; color:#9aa0a6; padding:6px 14px; border-radius:6px; cursor:pointer; font-size:12px; font-family:inherit; }}
  .band-controls button:hover {{ color:#c7cbd1; border-color:#4dabf7; }}
  .close-btn {{ float:right; background:none; border:none; color:#9aa0a6; font-size:20px; cursor:pointer; }}
  .chart-wrap {{ height:420px; position:relative; margin-top:8px; }}
</style>
</head>
<body>
  <a class="back" href="index.html">&larr; 홈</a>
  <h1>OP밴드 트래커</h1>
  <div class="updated">최종 갱신: {updated_at} &middot; 원본 파일 기준일 {src_mtime} &middot; {n_stocks}종목 &middot; 배수 = 시가총액 / 연간 영업이익 추정치(6월 기준 회계연도 스위칭)</div>

  <div class="filters">
    <label>검색 <input type="text" id="fSearch" placeholder="종목명/코드"></label>
    <label>섹터 <select id="fSector"><option value="">전체</option></select></label>
    <label>최신배수 최소 <input type="number" id="fMultMin" step="0.1"></label>
    <label>최신배수 최대 <input type="number" id="fMultMax" step="0.1"></label>
    <label><input type="checkbox" id="fIncludeNeg"> 적자(마이너스 배수) 포함</label>
    <label><input type="checkbox" id="fMinGap"> 최소 대비 근접(≤<input type="number" id="fMinGapPct" value="10" step="1" style="width:48px">%)만</label>
    <label>정렬
      <select id="fSort">
        <option value="latest_mult_asc">최신배수 낮은순</option>
        <option value="latest_mult_desc">최신배수 높은순</option>
        <option value="percentile_asc">과거 대비 저평가순(백분위 낮은순)</option>
        <option value="percentile_desc">과거 대비 고평가순(백분위 높은순)</option>
        <option value="min_gap_asc">최소 대비 근접순</option>
        <option value="mktcap_desc">시가총액 큰순</option>
        <option value="mktcap_asc">시가총액 작은순</option>
      </select>
    </label>
  </div>

  <table>
    <thead><tr>
      <th>종목명</th><th>코드</th><th>섹터</th><th>시가총액</th><th>최신배수</th><th>기간중 최소</th><th>기간중 최대</th><th>현재 백분위</th><th>최소 대비</th><th>기준일</th>
    </tr></thead>
    <tbody id="tbody"></tbody>
  </table>

  <div class="overlay" id="overlay">
    <div class="detail-card">
      <button class="close-btn" id="closeBtn">&times;</button>
      <h2 id="detailName"></h2>
      <div class="sub" id="detailSub"></div>
      <div class="range-bar" id="rangeBar"></div>
      <div class="custom-range-bar">
        <label>시작 <input type="date" id="rangeStart"></label>
        <label>종료 <input type="date" id="rangeEnd"></label>
        <button id="rangeApplyBtn">적용</button>
        <span class="custom-range-hint" id="rangeCustomHint"></span>
      </div>
      <div class="band-controls">
        <label>밴드 개수 <input type="number" id="bandCount" min="1" max="20" step="1"></label>
        <label>직접 배수 지정(콤마구분, 예: 5,10,20) <input type="text" id="bandCustom" placeholder="비워두면 자동"></label>
        <label>세로축 최대값(억원) <input type="number" id="yAxisMax" placeholder="자동"></label>
        <button id="bandApplyBtn">적용</button>
        <button id="bandResetBtn">자동으로</button>
      </div>
      <div class="chart-wrap"><canvas id="bandChart"></canvas></div>
    </div>
  </div>

<script>
const ROWS = {rows_json};
let chart = null;

function pctlClass(p) {{
  if (p <= 20) return 'pctl-low';
  if (p >= 80) return 'pctl-high';
  return '';
}}

// (최신배수 - 기간중 최소) / 최신배수 * 100 - 최신배수가 과거 최저치에 얼마나 근접했는지(%).
// 0%에 가까울수록 역대 최저 배수 근처, 마이너스 최신배수는 의미가 없어 null 처리.
function minGapPct(r) {{
  return r.latest_mult > 0 ? (r.latest_mult - r.hist_min_mult) / r.latest_mult * 100 : null;
}}

// 시가총액(원) -> "N조 M,MMM억원" 표시. 1조 미만이면 억원 단위만.
function fmtMktcap(won) {{
  if (won === null || won === undefined) return '-';
  const eok = won / 1e8;
  const jo = Math.floor(eok / 10000);
  const rest = Math.round(eok - jo * 10000);
  if (jo > 0) return `${{jo}}조 ${{rest.toLocaleString()}}억`;
  return `${{Math.round(eok).toLocaleString()}}억`;
}}

function applyFilters() {{
  const search = document.getElementById('fSearch').value.trim().toLowerCase();
  const sector = document.getElementById('fSector').value;
  const min = parseFloat(document.getElementById('fMultMin').value);
  const max = parseFloat(document.getElementById('fMultMax').value);
  const sort = document.getElementById('fSort').value;

  const includeNeg = document.getElementById('fIncludeNeg').checked;
  const minGapOn = document.getElementById('fMinGap').checked;
  const minGapThreshold = parseFloat(document.getElementById('fMinGapPct').value);
  let rows = ROWS.filter(r => {{
    if (!includeNeg && r.latest_mult <= 0) return false;
    if (search && !r.name.toLowerCase().includes(search) && !r.code.toLowerCase().includes(search)) return false;
    if (sector && r.sector !== sector) return false;
    if (!isNaN(min) && r.latest_mult < min) return false;
    if (!isNaN(max) && r.latest_mult > max) return false;
    if (minGapOn) {{
      const g = minGapPct(r);
      if (g === null || isNaN(minGapThreshold) || g > minGapThreshold) return false;
    }}
    return true;
  }});

  if (sort === 'min_gap_asc') {{
    rows.sort((a, b) => (minGapPct(a) ?? Infinity) - (minGapPct(b) ?? Infinity));
  }} else if (sort === 'mktcap_desc' || sort === 'mktcap_asc') {{
    const dir = sort === 'mktcap_asc' ? 1 : -1;
    rows.sort((a, b) => ((a.latest_mktcap ?? -Infinity) - (b.latest_mktcap ?? -Infinity)) * dir);
  }} else {{
    const [key, dir] = sort.includes('percentile') ? ['percentile', sort.endsWith('asc') ? 1 : -1]
      : ['latest_mult', sort.endsWith('asc') ? 1 : -1];
    rows.sort((a, b) => (a[key] - b[key]) * dir);
  }}

  const tbody = document.getElementById('tbody');
  tbody.innerHTML = rows.map(r => {{
    const g = minGapPct(r);
    return `
    <tr data-code="${{r.code}}">
      <td>${{r.name}}</td>
      <td>${{r.code}}</td>
      <td>${{r.sector ?? '-'}}</td>
      <td>${{fmtMktcap(r.latest_mktcap)}}</td>
      <td>${{r.latest_mult.toFixed(2)}}x</td>
      <td>${{r.hist_min_mult.toFixed(2)}}x</td>
      <td>${{r.hist_max_mult.toFixed(2)}}x</td>
      <td class="${{pctlClass(r.percentile)}}">${{r.percentile.toFixed(0)}}%ile</td>
      <td>${{g === null ? '-' : g.toFixed(1) + '%'}}</td>
      <td>${{r.latest_date}}</td>
    </tr>
  `;}}).join('');
  tbody.querySelectorAll('tr').forEach(tr => {{
    tr.addEventListener('click', () => openDetail(tr.dataset.code));
  }});
}}

const sectorSelect = document.getElementById('fSector');
[...new Set(ROWS.map(r => r.sector).filter(Boolean))].sort().forEach(s => {{
  const opt = document.createElement('option');
  opt.value = s;
  opt.textContent = s;
  sectorSelect.appendChild(opt);
}});

['fSearch', 'fSector', 'fMultMin', 'fMultMax', 'fSort', 'fIncludeNeg', 'fMinGap', 'fMinGapPct'].forEach(id => {{
  document.getElementById(id).addEventListener('input', applyFilters);
  document.getElementById(id).addEventListener('change', applyFilters);
}});

// 서버(build_op_band.py)의 pick_band_multiples()와 동일한 로직 - 밴드 개수를 사용자가
// 바꾸면 서버 재실행 없이 브라우저에서 즉시 다시 계산한다.
const NICE_STEPS = [1, 2, 5, 10, 15, 20, 25, 30, 50, 100, 200, 250, 500, 1000];

function pickBandMultiples(mults, targetLines) {{
  const positive = mults.filter(m => m != null && m > 0);
  if (!positive.length) return [];
  const end = Math.ceil(Math.max(...positive)) + 1;
  const threshold = end / targetLines;
  const step = NICE_STEPS.find(s => s >= threshold) ?? NICE_STEPS[NICE_STEPS.length - 1];
  const multiples = [];
  for (let v = step; v <= end; v += step) multiples.push(v);
  return multiples.length ? multiples : [step];
}}

let currentDetailData = null;
let currentBandMultiples = [];
let currentYMaxOverride = null;

// 기간 선택 - 볼린저밴드 트래커(build_bollinger_breakout.py)와 동일한 프리셋+커스텀 날짜 패턴.
const rangeBar = document.getElementById('rangeBar');
const rangeStartInput = document.getElementById('rangeStart');
const rangeEndInput = document.getElementById('rangeEnd');
const rangeCustomHint = document.getElementById('rangeCustomHint');
let currentRange = {{ mode: 'preset', days: null }};

const RANGE_OPTIONS = [
  {{ label: '1년', days: 365 }},
  {{ label: '3년', days: 1095 }},
  {{ label: '5년', days: 1825 }},
  {{ label: '전체', days: null }},
];

function computeRangeIndices(dates, range) {{
  if (!dates.length) return [0, -1];
  if (range.mode === 'custom') {{
    let startIdx = 0;
    if (range.start) {{
      const found = dates.findIndex(d => d >= range.start);
      startIdx = found < 0 ? dates.length : found;
    }}
    let endIdx = dates.length - 1;
    if (range.end) {{
      endIdx = -1;
      for (let i = dates.length - 1; i >= 0; i--) {{
        if (dates[i] <= range.end) {{ endIdx = i; break; }}
      }}
    }}
    return [startIdx, endIdx];
  }}
  let startIdx = 0;
  if (range.days !== null) {{
    const cutoff = new Date(dates[dates.length - 1]);
    cutoff.setDate(cutoff.getDate() - range.days);
    const found = dates.findIndex(d => new Date(d) >= cutoff);
    startIdx = found < 0 ? 0 : found;
  }}
  return [startIdx, dates.length - 1];
}}

function renderChart(data, bandMultiples, yMaxOverrideEok) {{
  currentBandMultiples = bandMultiples;
  currentYMaxOverride = yMaxOverrideEok ?? null;

  const [startIdx, endIdx] = computeRangeIndices(data.dates, currentRange);
  const dates = data.dates.slice(startIdx, endIdx + 1);
  const op = data.op.slice(startIdx, endIdx + 1);
  const mktcap = data.mktcap.slice(startIdx, endIdx + 1);

  document.getElementById('detailSub').textContent =
    `밴드선: ${{bandMultiples.length ? bandMultiples.map(m => m + 'x').join(', ') : '(표시할 배수 없음)'}}`;

  const datasets = bandMultiples.map((m, i) => ({{
    label: `${{m}}x`,
    data: op.map(v => v * m),
    borderColor: `hsl(${{200 + i * 30}}, 60%, 55%)`,
    backgroundColor: 'transparent',
    borderWidth: 1, borderDash: [4, 3], pointRadius: 0, tension: 0,
  }}));
  datasets.push({{
    label: '시가총액(실제)', data: mktcap,
    borderColor: '#e6e6e6', backgroundColor: 'transparent',
    borderWidth: 2.5, pointRadius: 0, tension: 0, order: 0,
  }});

  // 배수가 극단적으로 큰(FY2026/2027 원거리 컨센서스 이상치 등) 밴드선이 y축을 지배해서
  // 정작 봐야 할 실제 시가총액 선이 바닥에 눌려 안 보이는 문제 방지 - y축을 실제 시총
  // 범위에 맞춰 제한한다(밴드선은 그 위로 잘려 보이는 게 정상 - 그만큼 비싼 배수라는 뜻).
  // 자동 계산값이 여전히 보기 불편하면 사용자가 "세로축 최대값" 입력으로 직접 덮어쓸 수 있다.
  const mktcapVals = mktcap.filter(v => v != null);
  const maxMktcap = mktcapVals.length ? Math.max(...mktcapVals) : null;
  const minMktcap = mktcapVals.length ? Math.min(...mktcapVals) : null;
  const yMax = yMaxOverrideEok != null ? yMaxOverrideEok * 1e8 : (maxMktcap != null ? maxMktcap * 4.5 : undefined);
  const yMin = minMktcap != null && minMktcap < 0 ? minMktcap * 1.5 : 0;

  if (chart) chart.destroy();
  chart = new Chart(document.getElementById('bandChart').getContext('2d'), {{
    type: 'line',
    data: {{ labels: dates, datasets }},
    options: {{
      responsive: true, maintainAspectRatio: false,
      plugins: {{ legend: {{ labels: {{ color: '#e6e6e6', boxWidth: 14, font: {{ size: 11 }} }} }} }},
      scales: {{
        x: {{ ticks: {{ color: '#9aa0a6', maxTicksLimit: 10 }}, grid: {{ color: '#23262e' }} }},
        y: {{ min: yMin, max: yMax, ticks: {{ color: '#9aa0a6' }}, grid: {{ color: '#23262e' }} }},
      }}
    }}
  }});
}}

function applyCurrentRange() {{
  if (!currentDetailData) return;
  renderChart(currentDetailData, currentBandMultiples, currentYMaxOverride);
}}

function applyRange(days) {{
  currentRange = {{ mode: 'preset', days }};
  applyCurrentRange();
  document.querySelectorAll('.range-btn').forEach(btn => {{
    btn.classList.toggle('active', btn.dataset.days === String(days));
  }});
  rangeStartInput.value = '';
  rangeEndInput.value = '';
  rangeCustomHint.textContent = '';
}}

function applyCustomRange() {{
  const start = rangeStartInput.value || null;
  const end = rangeEndInput.value || null;
  if (!start && !end) return;
  if (start && end && start > end) {{
    rangeCustomHint.textContent = '시작일이 종료일보다 늦습니다.';
    return;
  }}
  currentRange = {{ mode: 'custom', start, end }};
  applyCurrentRange();
  document.querySelectorAll('.range-btn').forEach(btn => btn.classList.remove('active'));
  rangeCustomHint.textContent = `${{start || '처음'}} ~ ${{end || '최신'}} 구간 적용됨`;
}}

document.getElementById('rangeApplyBtn').addEventListener('click', applyCustomRange);
RANGE_OPTIONS.forEach(opt => {{
  const btn = document.createElement('button');
  btn.className = 'range-btn';
  btn.textContent = opt.label;
  btn.dataset.days = String(opt.days);
  btn.onclick = () => applyRange(opt.days);
  rangeBar.appendChild(btn);
}});

function applyBandControls() {{
  if (!currentDetailData) return;
  const customText = document.getElementById('bandCustom').value.trim();
  let bandMultiples;
  if (customText) {{
    bandMultiples = customText.split(',').map(s => parseFloat(s.trim())).filter(v => !isNaN(v) && v > 0).sort((a, b) => a - b);
  }} else {{
    const count = parseInt(document.getElementById('bandCount').value, 10) || TARGET_BAND_LINES_DEFAULT;
    bandMultiples = pickBandMultiples(currentDetailData.mult, count);
  }}
  const yMaxText = document.getElementById('yAxisMax').value.trim();
  const yMaxOverride = yMaxText ? parseFloat(yMaxText) : null;
  renderChart(currentDetailData, bandMultiples, isNaN(yMaxOverride) ? null : yMaxOverride);
}}

const TARGET_BAND_LINES_DEFAULT = 6;

function openDetail(code) {{
  fetch(`op_band_data/${{code}}.json`)
    .then(r => r.json())
    .then(data => {{
      currentDetailData = data;
      document.getElementById('detailName').textContent = `${{data.name}} (${{data.code}})`;
      const latestMult = data.mult[data.mult.length - 1];
      document.getElementById('detailName').textContent += latestMult != null ? ` - 최신 배수 ${{latestMult.toFixed(2)}}x` : '';
      document.getElementById('bandCount').value = data.bandMultiples.length || TARGET_BAND_LINES_DEFAULT;
      document.getElementById('bandCustom').value = '';
      document.getElementById('yAxisMax').value = '';
      if (data.dates.length) {{
        rangeStartInput.min = data.dates[0]; rangeStartInput.max = data.dates[data.dates.length - 1];
        rangeEndInput.min = data.dates[0]; rangeEndInput.max = data.dates[data.dates.length - 1];
      }}
      currentRange = {{ mode: 'preset', days: null }};
      rangeStartInput.value = '';
      rangeEndInput.value = '';
      rangeCustomHint.textContent = '';
      document.querySelectorAll('.range-btn').forEach(btn => btn.classList.toggle('active', btn.dataset.days === 'null'));
      renderChart(data, data.bandMultiples);
      document.getElementById('overlay').classList.add('open');
    }})
    .catch(() => alert('데이터를 불러올 수 없습니다.'));
}}

document.getElementById('bandApplyBtn').addEventListener('click', applyBandControls);
document.getElementById('bandResetBtn').addEventListener('click', () => {{
  if (!currentDetailData) return;
  document.getElementById('bandCustom').value = '';
  document.getElementById('yAxisMax').value = '';
  document.getElementById('bandCount').value = currentDetailData.bandMultiples.length || TARGET_BAND_LINES_DEFAULT;
  renderChart(currentDetailData, currentDetailData.bandMultiples);
}});

document.getElementById('closeBtn').addEventListener('click', () => {{
  document.getElementById('overlay').classList.remove('open');
}});
document.getElementById('overlay').addEventListener('click', (e) => {{
  if (e.target.id === 'overlay') document.getElementById('overlay').classList.remove('open');
}});

applyFilters();
</script>
</body>
</html>
"""


if __name__ == "__main__":
    main()
