"""
KOFIA FreeSIS에서 수동으로 다운로드한 엑셀을 data/kofia_raw.csv 로,
같은 파일에 포함된 투자자별 순매매대금은 data/krx_raw.csv 로 병합한다.

FreeSIS는 공식 API가 없고 (레거시 캔버스 렌더링이라 자동 스크래핑도 불가),
freesis.kofia.or.kr / kofia.or.kr 어디에도 자동조회 허용 여부를 명시한
이용약관 페이지가 없어 수동 다운로드 방식을 쓴다.

사용자가 FreeSIS "로우데이터" 다운로드로 받은 엑셀을 data/manual/ 에 넣으면
이 스크립트가 실행 시마다 다시 읽어서(파일이 갱신될 수 있으므로) 아래 두
CSV에 날짜 기준으로 upsert(기존 값 덮어쓰기) 한다.

기대하는 시트 레이아웃 ("로우데이터" 시트, 실제 다운로드 파일 기준):
  - A~H열: 구분(날짜), 투자자예탁금(파생상품 거래예수금 제외), 장내파생상품 거래예수금,
    대고객 RP매도잔고, 위탁매매 미수금, 반대매매금액, 반대매매비중(%), CMA잔고
  - J~N열: 일자, 기관합계, 기타법인, 개인, 외국인합계  ("거래대금 추이" = 투자자별 순매매대금)

시트가 없거나 헤더가 안 맞으면 해당 파일은 건너뛰고 경고만 출력한다
(예상과 다른 새 파일이 들어오면 추측해서 파싱하지 않고 알려주기 위함).
"""
import os

import pandas as pd

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
MANUAL_DIR = os.path.join(DATA_DIR, "manual")
KOFIA_OUT_PATH = os.path.join(DATA_DIR, "kofia_raw.csv")
KRX_OUT_PATH = os.path.join(DATA_DIR, "krx_raw.csv")

LEFT_HEADER = ["date", "investor_deposit", "deriv_deposit", "broker_rp_balance",
               "margin_call_unpaid", "margin_call_liquidation", "margin_liquidation_ratio", "cma_balance"]
RIGHT_HEADER = ["date", "inst_net_value", "other_corp_net_value", "indiv_net_value", "foreign_net_value"]


def parse_left_table(ws):
    rows = []
    r = 4
    while True:
        date_val = ws.cell(row=r, column=1).value
        if date_val is None:
            break
        rows.append([ws.cell(row=r, column=c).value for c in range(1, 9)])
        r += 1
    if not rows:
        return None
    df = pd.DataFrame(rows, columns=LEFT_HEADER)
    df["date"] = pd.to_datetime(df["date"], format="%Y/%m/%d", errors="coerce")
    return df.dropna(subset=["date"])


def parse_right_table(ws):
    rows = []
    r = 4
    while True:
        date_val = ws.cell(row=r, column=10).value
        if date_val is None:
            if r > 4 and ws.cell(row=r + 1, column=10).value is None and ws.cell(row=r + 2, column=10).value is None:
                break
            r += 1
            if r > ws.max_row:
                break
            continue
        rows.append([ws.cell(row=r, column=c).value for c in (10, 11, 12, 13, 14)])
        r += 1
    if not rows:
        return None
    df = pd.DataFrame(rows, columns=RIGHT_HEADER)
    df["date"] = pd.to_datetime(df["date"], format="%Y/%m/%d", errors="coerce")
    return df.dropna(subset=["date"])


def upsert(out_path, new_df):
    if new_df is None or new_df.empty:
        return 0
    if os.path.exists(out_path):
        existing = pd.read_csv(out_path, parse_dates=["date"])
        combined = pd.concat([existing, new_df], ignore_index=True)
    else:
        combined = new_df
    combined = combined.drop_duplicates(subset="date", keep="last").sort_values("date").reset_index(drop=True)
    os.makedirs(DATA_DIR, exist_ok=True)
    combined.to_csv(out_path, index=False, encoding="utf-8-sig")
    return len(combined)


def main():
    os.makedirs(MANUAL_DIR, exist_ok=True)
    files = [f for f in os.listdir(MANUAL_DIR) if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$")]

    if not files:
        print("data/manual/ 에 엑셀 파일이 없습니다.")
        return

    import openpyxl

    left_frames, right_frames = [], []
    for fname in files:
        path = os.path.join(MANUAL_DIR, fname)
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            print(f"  경고: {fname} 열기 실패 ({e}), 스킵")
            continue

        ws = wb["로우데이터"] if "로우데이터" in wb.sheetnames else wb.worksheets[0]

        left = parse_left_table(ws)
        right = parse_right_table(ws)

        if left is None and right is None:
            print(f"  경고: {fname}에서 예상한 레이아웃(로우데이터 시트, A/J열 날짜)을 못 찾음. 스킵")
            continue

        if left is not None:
            print(f"  {fname} [로우데이터 A~H열] -> kofia_raw.csv 후보 {len(left)}행 "
                  f"({left['date'].min().date()} ~ {left['date'].max().date()})")
            left_frames.append(left)
        if right is not None:
            print(f"  {fname} [로우데이터 J~N열, 거래대금추이] -> krx_raw.csv 후보 {len(right)}행 "
                  f"({right['date'].min().date()} ~ {right['date'].max().date()})")
            right_frames.append(right)

    kofia_rows = 0
    if left_frames:
        kofia_new = pd.concat(left_frames, ignore_index=True)
        kofia_rows = upsert(KOFIA_OUT_PATH, kofia_new)

    krx_rows = 0
    if right_frames:
        right_new = pd.concat(right_frames, ignore_index=True)
        if os.path.exists(KRX_OUT_PATH):
            existing_krx = pd.read_csv(KRX_OUT_PATH, parse_dates=["date"])
            merge_cols = [c for c in RIGHT_HEADER if c != "date"]
            existing_krx = existing_krx.set_index("date")
            right_new_idx = right_new.set_index("date")
            union_index = existing_krx.index.union(right_new_idx.index)
            existing_krx = existing_krx.reindex(union_index)
            for col in merge_cols:
                if col not in existing_krx.columns:
                    existing_krx[col] = pd.NA
                existing_krx[col] = right_new_idx[col].reindex(union_index).combine_first(existing_krx[col])
            existing_krx.index.name = "date"
            combined_krx = existing_krx.reset_index().sort_values("date")
        else:
            combined_krx = right_new.sort_values("date")
        combined_krx.to_csv(KRX_OUT_PATH, index=False, encoding="utf-8-sig")
        krx_rows = len(combined_krx)

    print(f"\nkofia_raw.csv: {kofia_rows}행" if kofia_rows else "\nkofia_raw.csv: 변경 없음")
    print(f"krx_raw.csv (투자자별 순매매대금 컬럼 갱신): {krx_rows}행" if krx_rows else "krx_raw.csv: 변경 없음")


if __name__ == "__main__":
    main()
