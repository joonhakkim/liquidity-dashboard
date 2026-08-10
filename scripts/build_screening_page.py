"""
data/screening/*.csv 를 모아 docs/screening.html (주식 스크리닝 페이지)를 만든다.

입력:
- op_growth_screen.csv: 시총2000억+/영업이익 2026->2027 컨센서스 50%+ 성장 종목 목록
- per_band.csv: 최근 10개년 연말 스냅샷 기준 PER 밴드 근사치
- dart_annual_trend.csv: DART 사업보고서 기준 최근 6개년 매출액/영업이익 실적(회사별 시계열)
- manual/*데이터 모음*.xlsm '섹터별 구성 종목' 시트: 종목 -> 섹터 매핑

결측 데이터는 그대로 null로 둬서 페이지에서 'N/A'로 표시한다(값을 지어내지 않는다).
"""
import glob
import json
import os
import warnings

import openpyxl
import pandas as pd

from dart_client import load_corp_code_map as load_corp_code_map_safe

warnings.filterwarnings("ignore")

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
SCREEN_DIR = os.path.join(DATA_DIR, "screening")
MANUAL_DIR = os.path.join(DATA_DIR, "manual")
DOCS_DIR = os.path.join(os.path.dirname(__file__), "..", "docs")
OUT_PATH = os.path.join(DOCS_DIR, "screening.html")


def find_workbook():
    candidates = glob.glob(os.path.join(MANUAL_DIR, "*데이터 모음*.xls*"))
    candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]
    return max(candidates, key=os.path.getmtime) if candidates else None


def load_sector_map():
    path = find_workbook()
    if not path:
        return {}
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    if "섹터별 구성 종목" not in wb.sheetnames:
        return {}
    ws = wb["섹터별 구성 종목"]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r[0] == "섹터")
    sectors = rows[header_idx]
    mapping = {}
    for r in rows[header_idx + 1:]:
        for col, sector in enumerate(sectors):
            if col == 0 or not sector:
                continue
            name = r[col] if col < len(r) else None
            if name and name not in mapping:
                mapping[name] = sector
    return mapping


def nz(v):
    """NaN -> None (JSON에 안전하게)."""
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    return v


def quarter_sort_key(row):
    return (row["연도"], {"Q1": 1, "Q2": 2, "Q3": 3, "Q4": 4}[row["분기"]])


def main():
    screen = pd.read_csv(os.path.join(SCREEN_DIR, "op_growth_screen.csv"))
    per = pd.read_csv(os.path.join(SCREEN_DIR, "per_band.csv"))
    quarterly = pd.read_csv(os.path.join(SCREEN_DIR, "dart_quarterly.csv"))
    sector_map = load_sector_map()

    prelim_path = os.path.join(SCREEN_DIR, "dart_preliminary_q2.csv")
    prelim_by_name = {}
    if os.path.exists(prelim_path):
        prelim_df = pd.read_csv(prelim_path)
        prelim_by_name = {row["종목명"]: row for _, row in prelim_df.iterrows()} if not prelim_df.empty else {}

    # 수주잔고: DART 정기보고서 본문에서 회사마다 다른 표기("기말수주잔고"/"건설계약 수주잔고"/
    # "공사계약 분기말잔액" 등)로 흩어져있는 걸 휴리스틱으로 추출한 것 - 조선/건설/방산처럼
    # 수주 기반 업종만 있고, 대부분 기업은 이 공시 자체가 없다(자동으로 표시 안 됨).
    backlog_history_path = os.path.join(SCREEN_DIR, "order_backlog_history.csv")
    backlog_by_name = {}
    if os.path.exists(backlog_history_path):
        bh = pd.read_csv(backlog_history_path)
        bh["기준일"] = pd.to_datetime(bh["기준일"], format="%Y%m%d")
        for name, g in bh.groupby("종목명"):
            g = g.sort_values("기준일")
            backlog_by_name[name] = {
                "labels": g["기준일"].dt.strftime("%Y-%m-%d").tolist(),
                "values": [nz(v) for v in g["수주잔고(억원)"].tolist()],
            }

    quarterly_by_name = {}
    q2_yoy_by_name = {}
    for name, g in quarterly.groupby("종목명"):
        g = g.assign(_k=g.apply(quarter_sort_key, axis=1)).sort_values("_k")
        quarterly_by_name[name] = {
            "labels": [f"{int(y)} {q}" for y, q in zip(g["연도"], g["분기"])],
            "revenue": [nz(v) for v in g["매출액"].tolist()],
            "op": [nz(v) for v in g["영업이익"].tolist()],
            "margin": [nz(v) for v in g["영업이익률"].tolist()],
        }
        this_q2 = g[(g["연도"] == 2026) & (g["분기"] == "Q2")]
        prev_q2 = g[(g["연도"] == 2025) & (g["분기"] == "Q2")]
        rev_yoy = op_yoy = None
        if not this_q2.empty and not prev_q2.empty:
            rt, rp = this_q2.iloc[0]["매출액"], prev_q2.iloc[0]["매출액"]
            ot, op_ = this_q2.iloc[0]["영업이익"], prev_q2.iloc[0]["영업이익"]
            if pd.notna(rt) and pd.notna(rp) and rp != 0:
                rev_yoy = rt / rp - 1
            if pd.notna(ot) and pd.notna(op_) and op_ != 0:
                op_yoy = ot / op_ - 1
        q2_yoy_by_name[name] = (nz(rev_yoy), nz(op_yoy))

        # 2026 Q2 정식 분기 데이터가 아직 없으면(반기보고서 마감 전) 잠정실적으로 채워서
        # 상세화면 분기 막대차트에 최소한 잠정치라도 보이게 한다.
        labels = quarterly_by_name[name]["labels"]
        prelim = prelim_by_name.get(name)
        if "2026 Q2" not in labels and isinstance(prelim, pd.Series) and prelim.get("상태") == "ok":
            rev_eok, op_eok = prelim.get("매출액_당기(억원)"), prelim.get("영업이익_당기(억원)")
            rev_won = rev_eok * 1e8 if pd.notna(rev_eok) else None
            op_won = op_eok * 1e8 if pd.notna(op_eok) else None
            if rev_won is not None or op_won is not None:
                margin = (op_won / rev_won) if (op_won is not None and rev_won not in (None, 0)) else None
                quarterly_by_name[name]["labels"].append("2026 Q2(잠정)")
                quarterly_by_name[name]["revenue"].append(rev_won)
                quarterly_by_name[name]["op"].append(op_won)
                quarterly_by_name[name]["margin"].append(margin)
                quarterly_by_name[name]["prelimIndex"] = len(quarterly_by_name[name]["labels"]) - 1

    per_by_name = {row["종목명"]: row for _, row in per.iterrows()} if not per.empty else {}
    pbr_path = os.path.join(SCREEN_DIR, "pbr_band.csv")
    pbr_by_name = {}
    if os.path.exists(pbr_path):
        pbr_df = pd.read_csv(pbr_path)
        pbr_by_name = {row["종목명"]: row for _, row in pbr_df.iterrows()} if not pbr_df.empty else {}

    _, _, name_to_stock_code = load_corp_code_map_safe()

    companies = []
    for _, row in screen.iterrows():
        name = row["종목명"]
        p = per_by_name.get(name, {})
        b = pbr_by_name.get(name, {})
        prelim = prelim_by_name.get(name, {})
        rev_yoy, op_yoy = q2_yoy_by_name.get(name, (None, None))
        q2_source = "quarterly_filing" if (rev_yoy is not None or op_yoy is not None) else None
        if isinstance(prelim, pd.Series) and prelim.get("상태") == "ok":
            prelim_rev_yoy = nz(prelim.get("매출액_YoY(%)"))
            prelim_op_yoy = nz(prelim.get("영업이익_YoY(%)"))
            if prelim_rev_yoy is not None:
                rev_yoy = prelim_rev_yoy / 100
                q2_source = "preliminary"
            if prelim_op_yoy is not None:
                op_yoy = prelim_op_yoy / 100
                q2_source = "preliminary"
        companies.append({
            "name": name,
            "stockCode": name_to_stock_code.get(name),
            "sector": sector_map.get(name),
            "marketCap": nz(row["시가총액(억원)"]),
            "op2026": nz(row["2026_영업이익(십억원)"]),
            "op2027": nz(row["2027_영업이익(십억원)"]),
            "opGrowth": nz(row["영업이익_증가율"]),
            "q2RevYoy": rev_yoy,
            "q2OpYoy": op_yoy,
            "q2Source": q2_source,
            "currentPer": nz(p.get("현재PER")),
            # 적자가 지속돼 흑자전환이 안 돼 PER 자체를 계산 못하는 경우는 N/A로 비우면
            # "PER 밴드위치 필터"에서 그냥 빠져버린다 - 밴드 최하단(0%ile)으로 취급해 필터에 남긴다.
            "perPercentile": 0 if isinstance(p, pd.Series) and p.get("상태") == "no_positive_ttm_eps" else nz(p.get("밴드내_위치_percentile")),
            "currentPbr": nz(b.get("현재PBR")),
            "pbrPercentile": 0 if isinstance(b, pd.Series) and b.get("상태") == "no_positive_bps" else nz(b.get("밴드내_위치_percentile")),
            "quarterly": quarterly_by_name.get(name, {"labels": [], "revenue": [], "op": [], "margin": []}),
            "orderBacklog": backlog_by_name.get(name),
        })

    companies.sort(key=lambda c: (c["opGrowth"] is None, -(c["opGrowth"] or 0)))

    html = TEMPLATE.replace("__COMPANIES_JSON__", json.dumps(companies, ensure_ascii=False))
    os.makedirs(DOCS_DIR, exist_ok=True)
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"저장 완료: {OUT_PATH} ({len(companies)}개 종목)")


TEMPLATE = """<!doctype html>
<html lang="ko">
<head>
<meta charset="utf-8">
<title>주식 스크리닝</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
  body { font-family: -apple-system, "Malgun Gothic", sans-serif; background:#0f1115; color:#e6e6e6; margin:0; padding:24px; }
  h1 { font-size:20px; margin:0 0 4px 0; }
  a.back { color:#4dabf7; font-size:13px; text-decoration:none; }
  .note { color:#9aa0a6; font-size:12px; margin:10px 0 20px 0; line-height:1.6; }
  .note b { color:#ffa94d; }
  table { width:100%; border-collapse:collapse; font-size:13px; }
  th, td { padding:8px 10px; text-align:right; border-bottom:1px solid #23262e; white-space:nowrap; }
  th:first-child, td:first-child { text-align:left; }
  th { color:#9aa0a6; font-weight:normal; cursor:pointer; position:sticky; top:0; background:#0f1115; }
  th:hover { color:#4dabf7; }
  tbody tr { cursor:pointer; }
  tbody tr:hover { background:#1a1d24; }
  .pos { color:#63e6be; }
  .neg { color:#ff6b6b; }
  .muted { color:#5a616e; }
  .table-wrap { max-height:calc(100vh - 160px); overflow:auto; border:1px solid #23262e; border-radius:8px; }
  #overlay { display:none; position:fixed; inset:0; background:rgba(0,0,0,0.6); z-index:10; }
  #overlay.open { display:flex; align-items:flex-start; justify-content:center; padding:40px 20px; overflow:auto; }
  #detail { background:#1a1d24; border-radius:12px; padding:24px; width:100%; max-width:820px; }
  #detail h2 { margin:0 0 4px 0; font-size:18px; }
  #detail .sector { color:#9aa0a6; font-size:13px; margin-bottom:16px; }
  .metric-grid { display:grid; grid-template-columns: repeat(auto-fit, minmax(150px, 1fr)); gap:12px; margin-bottom:20px; }
  .metric { background:#12141a; border-radius:8px; padding:10px 12px; }
  .metric .label { color:#9aa0a6; font-size:11px; }
  .metric .value { font-size:16px; font-weight:bold; margin-top:2px; }
  .chart-wrap { height:260px; position:relative; margin-bottom:20px; }
  .per-band-wrap { margin-bottom:16px; }
  .per-band-title { color:#9aa0a6; font-size:12px; margin-bottom:6px; }
  .per-band-chart-wrap { height:280px; position:relative; }
  #closeBtn { float:right; background:none; border:none; color:#9aa0a6; font-size:20px; cursor:pointer; }
  .band-range-bar { display:flex; gap:6px; margin-bottom:14px; }
  .range-btn { background:#1a1d24; border:1px solid #2a2e37; color:#9aa0a6; padding:5px 12px;
    border-radius:999px; cursor:pointer; font-size:12px; font-family:inherit; }
  .range-btn:hover { color:#c7cbd1; border-color:#4dabf7; }
  .range-btn.active { background:#4dabf7; color:#0f1115; border-color:#4dabf7; font-weight:bold; }
  .filter-bar { display:flex; flex-wrap:wrap; align-items:center; gap:16px; margin-bottom:16px; padding:12px 16px;
    background:#1a1d24; border:1px solid #23262e; border-radius:10px; font-size:13px; color:#c7cbd1; }
  .filter-bar label { display:flex; align-items:center; gap:6px; white-space:nowrap; }
  .filter-bar input[type="number"] { width:70px; background:#12141a; border:1px solid #2a2e37; color:#e6e6e6;
    border-radius:6px; padding:5px 8px; font-size:13px; }
  .filter-bar input[type="checkbox"] { width:14px; height:14px; }
  .filter-bar input[type="text"] { width:140px; background:#12141a; border:1px solid #2a2e37; color:#e6e6e6;
    border-radius:6px; padding:5px 8px; font-size:13px; }
</style>
</head>
<body>
  <a class="back" href="index.html">&larr; 홈</a> &middot; <a class="back" href="liquidity.html">유동성 대시보드 &rarr;</a>
  <h1>주식 스크리닝</h1>
  <div class="note">
    기본 모집단: 시가총액이 있는 전체 상장사(스팩·ETF 제외, DART 일일 API 호출 한도 때문에 분기실적/PER밴드는
    며칠에 걸쳐 순차적으로 채워짐). 성장률 컷은 없고 아래 필터로 직접 좁혀서 보세요.
    <b>매출액 컨센서스는 원본 데이터에 없어 영업이익만으로 스크리닝했습니다.</b>
    2분기 실적 YoY는 DART 잠정실적(공정공시) 공시가 있으면 그걸 우선 쓰고, 없으면 반기보고서 제출 마감(8/14) 전이라 결측입니다.
    PER·PBR 밴드는 분기별 TTM EPS/BPS와 월별 종가로 계산한 <b>근사치</b>이며 정식 리서치 밴드차트와는 값이 다를 수 있습니다.
    사업부별 매출비중·제품가격 추이는 자동으로 가져올 무료 소스가 없어 이번 페이지에는 포함하지 않았습니다.
  </div>
  <div class="filter-bar">
    <label>종목명 검색 <input type="text" id="fSearch" placeholder="예: 삼성전자"></label>
    <label>OP 증가율 &ge; <input type="number" id="fOpGrowth" placeholder="예: 30" step="5">%</label>
    <label>시가총액 &ge; <input type="number" id="fMarketCap" placeholder="예: 5000">억</label>
    <label>PER 밴드위치 &le; <input type="number" id="fPerPct" placeholder="예: 30" min="0" max="100">%ile <span class="muted">(낮을수록 밴드 하단)</span></label>
    <label><input type="checkbox" id="fHasPer"> PER 데이터 있는 종목만</label>
    <button id="fApply" class="range-btn">필터 적용</button>
    <button id="fReset" class="range-btn">초기화</button>
    <span id="fCount" class="muted"></span>
  </div>
  <div class="table-wrap">
  <table id="tbl">
    <thead><tr>
      <th data-key="name">종목명</th>
      <th data-key="sector">섹터</th>
      <th data-key="marketCap">시가총액(억원)</th>
      <th data-key="op2026">2026 OP(십억)</th>
      <th data-key="op2027">2027 OP(십억)</th>
      <th data-key="opGrowth">OP 증가율</th>
      <th data-key="q2RevYoy">2Q 매출YoY</th>
      <th data-key="q2OpYoy">2Q 영업이익YoY</th>
      <th data-key="currentPer">현재PER</th>
      <th data-key="perPercentile">PER밴드위치</th>
      <th data-key="currentPbr">현재PBR</th>
      <th data-key="pbrPercentile">PBR밴드위치</th>
    </tr></thead>
    <tbody id="tbody"></tbody>
  </table>
  </div>

  <div id="overlay">
    <div id="detail"></div>
  </div>

<script>
const COMPANIES = __COMPANIES_JSON__;
let sortKey = 'opGrowth', sortDir = -1;
let filters = { search: null, opGrowth: null, marketCap: null, perPct: null, hasPer: false };

function pct(v) { return v === null || v === undefined ? '<span class="muted">N/A</span>' : (v*100).toFixed(1) + '%'; }
function num(v, digits) { return v === null || v === undefined ? '<span class="muted">N/A</span>' : Number(v).toLocaleString(undefined, {maximumFractionDigits: digits ?? 0}); }
function colorize(v, html) { if (v === null || v === undefined) return html; return v >= 0 ? '<span class="pos">'+html+'</span>' : '<span class="neg">'+html+'</span>'; }

function applyFilters(list) {
  return list.filter(c => {
    if (filters.search && !c.name.toLowerCase().includes(filters.search)) return false;
    if (filters.opGrowth !== null && (c.opGrowth === null || c.opGrowth * 100 < filters.opGrowth)) return false;
    if (filters.marketCap !== null && (c.marketCap === null || c.marketCap < filters.marketCap)) return false;
    if (filters.perPct !== null && (c.perPercentile === null || c.perPercentile > filters.perPct)) return false;
    if (filters.hasPer && c.currentPer === null) return false;
    return true;
  });
}

function readFilterInputs() {
  const v = id => { const el = document.getElementById(id); const raw = el.value.trim(); return raw === '' ? null : Number(raw); };
  const searchRaw = document.getElementById('fSearch').value.trim();
  filters = {
    search: searchRaw === '' ? null : searchRaw.toLowerCase(),
    opGrowth: v('fOpGrowth'),
    marketCap: v('fMarketCap'),
    perPct: v('fPerPct'),
    hasPer: document.getElementById('fHasPer').checked,
  };
}

function render() {
  const filtered = applyFilters(COMPANIES);
  document.getElementById('fCount').textContent = `${filtered.length} / ${COMPANIES.length}개 종목`;
  const rows = filtered.slice().sort((a,b) => {
    const av = a[sortKey], bv = b[sortKey];
    if (av === null || av === undefined) return 1;
    if (bv === null || bv === undefined) return -1;
    if (typeof av === 'string') return sortDir * av.localeCompare(bv);
    return sortDir * (av - bv);
  });
  const tbody = document.getElementById('tbody');
  tbody.innerHTML = rows.map((c, i) => `
    <tr data-idx="${COMPANIES.indexOf(c)}">
      <td>${c.name}</td>
      <td class="muted">${c.sector ?? ''}</td>
      <td>${num(c.marketCap)}</td>
      <td>${num(c.op2026, 1)}</td>
      <td>${num(c.op2027, 1)}</td>
      <td>${colorize(c.opGrowth, pct(c.opGrowth))}</td>
      <td>${colorize(c.q2RevYoy, pct(c.q2RevYoy))}${c.q2Source === 'preliminary' ? ' <span class="muted" title="DART 잠정실적">잠정</span>' : ''}</td>
      <td>${colorize(c.q2OpYoy, pct(c.q2OpYoy))}</td>
      <td>${c.currentPer === null && c.perPercentile === 0 ? '<span class="muted" title="적자 지속(흑자전환 안됨)">적자</span>' : num(c.currentPer, 1)}</td>
      <td>${c.perPercentile === null || c.perPercentile === undefined ? '<span class="muted">N/A</span>' : c.perPercentile.toFixed(0) + '%ile'}</td>
      <td>${c.currentPbr === null && c.pbrPercentile === 0 ? '<span class="muted">N/A</span>' : num(c.currentPbr, 2)}</td>
      <td>${c.pbrPercentile === null || c.pbrPercentile === undefined ? '<span class="muted">N/A</span>' : c.pbrPercentile.toFixed(0) + '%ile'}</td>
    </tr>
  `).join('');
  tbody.querySelectorAll('tr').forEach(tr => {
    tr.addEventListener('click', () => openDetail(COMPANIES[Number(tr.dataset.idx)]));
  });
}

document.querySelectorAll('th[data-key]').forEach(th => {
  th.addEventListener('click', () => {
    const key = th.dataset.key;
    sortDir = (sortKey === key) ? -sortDir : -1;
    sortKey = key;
    render();
  });
});

document.getElementById('fApply').addEventListener('click', () => { readFilterInputs(); render(); });
document.getElementById('fReset').addEventListener('click', () => {
  document.getElementById('fSearch').value = '';
  ['fOpGrowth', 'fMarketCap', 'fPerPct'].forEach(id => document.getElementById(id).value = '');
  document.getElementById('fHasPer').checked = false;
  filters = { search: null, opGrowth: null, marketCap: null, perPct: null, hasPer: false };
  render();
});
document.getElementById('fSearch').addEventListener('input', () => { readFilterInputs(); render(); });
['fOpGrowth', 'fMarketCap', 'fPerPct'].forEach(id => {
  document.getElementById(id).addEventListener('keydown', e => { if (e.key === 'Enter') { readFilterInputs(); render(); } });
});

let detailChart = null;
function openDetail(c) {
  const overlay = document.getElementById('overlay');
  const detail = document.getElementById('detail');
  const perYears = Object.keys(c.perByYear || {}).sort();
  detail.innerHTML = `
    <button id="closeBtn">&times;</button>
    <h2>${c.name}</h2>
    <div class="sector">${c.sector ?? '섹터 미분류'}</div>
    <div class="metric-grid">
      <div class="metric"><div class="label">시가총액</div><div class="value">${num(c.marketCap)}억</div></div>
      <div class="metric"><div class="label">2026 OP(컨센서스)</div><div class="value">${num(c.op2026,1)}십억</div></div>
      <div class="metric"><div class="label">2027 OP(컨센서스)</div><div class="value">${num(c.op2027,1)}십억</div></div>
      <div class="metric"><div class="label">OP 증가율</div><div class="value">${pct(c.opGrowth)}</div></div>
      <div class="metric"><div class="label">2Q 매출 YoY${c.q2Source === 'preliminary' ? ' (잠정실적)' : ''}</div><div class="value">${pct(c.q2RevYoy)}</div></div>
      <div class="metric"><div class="label">2Q 영업이익 YoY${c.q2Source === 'preliminary' ? ' (잠정실적)' : ''}</div><div class="value">${pct(c.q2OpYoy)}</div></div>
      <div class="metric"><div class="label">현재 PER / PBR</div><div class="value">${c.currentPer !== null ? c.currentPer.toFixed(1) : 'N/A'} / ${c.currentPbr !== null ? c.currentPbr.toFixed(2) : 'N/A'}</div></div>
    </div>
    <div class="chart-wrap"><canvas id="trendChart"></canvas></div>
    <div id="backlogSection"></div>
    <div class="band-range-bar" id="bandRangeBar"></div>
    <div class="per-band-wrap">
      <div class="per-band-title">PER 밴드 (주가선 + TTM EPS&times;PER배수 밴드선, 확정 공시된 실적만 사용)</div>
      <div class="per-band-chart-wrap"><canvas id="perBandChart"></canvas></div>
    </div>
    <div class="per-band-wrap">
      <div class="per-band-title">PBR 밴드 (주가선 + BPS&times;PBR배수 밴드선)</div>
      <div class="per-band-chart-wrap"><canvas id="pbrBandChart"></canvas></div>
    </div>
  `;
  document.getElementById('closeBtn').addEventListener('click', closeDetail);
  overlay.classList.add('open');

  if (detailChart) detailChart.destroy();
  const ctx = document.getElementById('trendChart').getContext('2d');
  const q = c.quarterly || {labels: [], revenue: [], op: [], margin: []};
  const barColor = (base, dim) => q.labels.map((_, i) => i === q.prelimIndex ? dim : base);
  detailChart = new Chart(ctx, {
    data: {
      labels: q.labels,
      datasets: [
        { type: 'line', label: '영업이익률(%)', data: q.margin.map(v => v === null ? null : v * 100),
          borderColor: '#ffa94d', backgroundColor: 'transparent', yAxisID: 'y1', tension: 0.2, order: 0 },
        { type: 'bar', label: '매출액(원)', data: q.revenue, backgroundColor: barColor('#4dabf7', 'rgba(77,171,247,0.35)'), yAxisID: 'y', order: 1 },
        { type: 'bar', label: '영업이익(원)', data: q.op, backgroundColor: barColor('#63e6be', 'rgba(99,230,190,0.35)'), yAxisID: 'y', order: 1 },
      ]
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: {
        legend: { labels: { color: '#e6e6e6' } },
        tooltip: { callbacks: { footer: items => items[0].dataIndex === q.prelimIndex ? '⚠ DART 잠정실적(공정공시) 기준, 확정치와 다를 수 있음' : '' } },
      },
      scales: {
        x: { ticks: { color: '#9aa0a6' }, grid: { color: '#23262e' } },
        y: { position: 'left', ticks: { color: '#9aa0a6' }, grid: { color: '#23262e' } },
        y1: { position: 'right', ticks: { color: '#ffa94d', callback: v => v + '%' }, grid: { drawOnChartArea: false } },
      }
    }
  });

  renderBacklogSection(c);
  loadBandCharts(c);
}

let backlogChart = null;
function renderBacklogSection(c) {
  const el = document.getElementById('backlogSection');
  if (!c.orderBacklog || c.orderBacklog.labels.length === 0) { el.innerHTML = ''; return; }
  el.innerHTML = `
    <div class="per-band-wrap">
      <div class="per-band-title">수주잔고(억원) - DART 정기보고서 본문에서 자동 추출한 근사치, 조선·건설 등 수주 기반 업종만 있음</div>
      <div class="chart-wrap"><canvas id="backlogChart"></canvas></div>
    </div>
  `;
  if (backlogChart) backlogChart.destroy();
  const ctx = document.getElementById('backlogChart').getContext('2d');
  backlogChart = new Chart(ctx, {
    type: 'bar',
    data: {
      labels: c.orderBacklog.labels,
      datasets: [{ label: '수주잔고(억원)', data: c.orderBacklog.values, backgroundColor: '#b197fc' }],
    },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: { legend: { display: false } },
      scales: {
        x: { ticks: { color: '#9aa0a6' }, grid: { color: '#23262e' } },
        y: { ticks: { color: '#9aa0a6' }, grid: { color: '#23262e' } },
      }
    }
  });
}

const BAND_RANGES = [
  { label: '1년', months: 12 },
  { label: '3년', months: 36 },
  { label: '5년', months: 60 },
  { label: '전체', months: null },
];
let bandRangeMonths = 36;
let perBandChart = null, pbrBandChart = null;
let currentBandData = null;

function sliceByRange(points) {
  if (bandRangeMonths === null) return points;
  return points.slice(-bandRangeMonths);
}

function renderBandChart(canvasId, multiples, bands, price, existingChart) {
  if (existingChart) existingChart.destroy();
  const canvasEl = document.getElementById(canvasId);
  if (!canvasEl) return null;  // 다른 종목 클릭 등으로 캔버스가 이미 교체/제거된 경우
  const ctx = canvasEl.getContext('2d');
  const pricePts = sliceByRange(price);
  const labels = pricePts.map(p => p[0]);
  const bandColors = ['#5c5f66', '#748ffc', '#4dabf7', '#63e6be', '#ffd43b'];
  const datasets = multiples.map((m, i) => ({
    label: `${m}x`,
    data: sliceByRange(bands[`x${m}`]).map(p => p[1]),
    borderColor: bandColors[i % bandColors.length],
    backgroundColor: 'transparent',
    borderDash: [4, 3],
    pointRadius: 0,
    borderWidth: 1,
  }));
  datasets.push({
    label: '주가',
    data: pricePts.map(p => p[1]),
    borderColor: '#ff6b6b',
    backgroundColor: 'transparent',
    pointRadius: 0,
    borderWidth: 2,
  });
  return new Chart(ctx, {
    type: 'line',
    data: { labels, datasets },
    options: {
      responsive: true, maintainAspectRatio: false,
      plugins: { legend: { labels: { color: '#e6e6e6', boxWidth: 12 } } },
      scales: {
        x: { ticks: { color: '#9aa0a6', maxTicksLimit: 10 }, grid: { color: '#23262e' } },
        y: { ticks: { color: '#9aa0a6' }, grid: { color: '#23262e' } },
      }
    }
  });
}

function renderBandRangeBar() {
  const bar = document.getElementById('bandRangeBar');
  bar.innerHTML = BAND_RANGES.map(r =>
    `<button class="range-btn ${r.months === bandRangeMonths ? 'active' : ''}" data-months="${r.months}">${r.label}</button>`
  ).join('');
  bar.querySelectorAll('button').forEach(btn => {
    btn.addEventListener('click', () => {
      bandRangeMonths = btn.dataset.months === 'null' ? null : Number(btn.dataset.months);
      renderBandRangeBar();
      redrawBandCharts();
    });
  });
}

function redrawBandCharts() {
  if (!currentBandData) return;
  const data = currentBandData;
  if (data.perMultiples) {
    perBandChart = renderBandChart('perBandChart', data.perMultiples, data.perBands, data.price, perBandChart);
  }
  if (data.pbrMultiples) {
    pbrBandChart = renderBandChart('pbrBandChart', data.pbrMultiples, data.pbrBands, data.price, pbrBandChart);
  }
}

function loadBandCharts(c) {
  currentBandData = null;
  renderBandRangeBar();
  const perWrap = document.querySelectorAll('.per-band-chart-wrap')[0];
  const pbrWrap = document.querySelectorAll('.per-band-chart-wrap')[1];
  if (!c.stockCode) {
    perWrap.innerHTML = pbrWrap.innerHTML = '<div class="muted" style="padding:20px 0;">종목코드 없음</div>';
    return;
  }
  fetch(`screening_data/${c.stockCode}.json`)
    .then(r => { if (!r.ok) throw new Error('no data'); return r.json(); })
    .then(data => {
      currentBandData = data;
      if (!data.perMultiples) perWrap.innerHTML = '<div class="muted" style="padding:20px 0;">PER 밴드 데이터 부족(흑자 이력 없음 등)으로 표시할 수 없습니다.</div>';
      if (!data.pbrMultiples) pbrWrap.innerHTML = '<div class="muted" style="padding:20px 0;">PBR 밴드 데이터 부족으로 표시할 수 없습니다.</div>';
      redrawBandCharts();
    })
    .catch(() => {
      perWrap.innerHTML = '<div class="muted" style="padding:20px 0;">밴드 데이터를 불러올 수 없습니다.</div>';
      pbrWrap.innerHTML = '';
    });
}

function closeDetail() {
  document.getElementById('overlay').classList.remove('open');
}
document.getElementById('overlay').addEventListener('click', (e) => {
  if (e.target.id === 'overlay') closeDetail();
});

render();
</script>
</body>
</html>
"""

if __name__ == "__main__":
    main()
