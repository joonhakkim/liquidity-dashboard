"""
사용자가 데이터 터미널(FnGuide/DataGuide류)에서 내려받은 "수급정리" 엑셀을
data/manual/ 에서 찾아 파싱하고, 종목별 개인/기관/외국인 순매수를 KRX Open API로
확인한 코스피/코스닥 소속에 따라 나눠 합산해서 data/investor_flow_raw.csv 로 저장한다.

기대하는 파일: data/manual/수급정리*.xlsm (파일명에 "수급정리"가 들어가면 인식)
  - 시트 "개인_순매수", "기관_순매수", "외국인_순매수": 1행=종목명 헤더(B열부터),
    A열=날짜, 나머지 셀=해당 종목의 그날 순매수금액(백만원 단위로 추정)
  - 시트 "원본데이터": 8행=Code(예: A005930), 9행=Name, 3열 간격으로 종목 반복
    (개인/기관/외국인 3개 컬럼이 종목당 세트라서 3칸씩 건너뜀)

코스피/코스닥 분류는 이 파일 안에 없어서, 이미 승인된 KRX Open API
(유가증권/코스닥 일별매매정보, stk_bydd_trd / ksq_bydd_trd)로 그날 상장된
전종목 코드를 받아 대조한다 (로그인 불필요, 이미 다른 스크립트에서도 사용 중).

외국인 보유비중(%)은 이 파일에도, 지금까지 찾은 어떤 무료 소스에도 없어서
여전히 비워둔다.
"""
import glob
import os
from datetime import datetime, timedelta

import pandas as pd
import requests
from dotenv import load_dotenv

load_dotenv()

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
MANUAL_DIR = os.path.join(DATA_DIR, "manual")
OUT_PATH = os.path.join(DATA_DIR, "investor_flow_raw.csv")

KRX_OPEN_API_KEY = os.environ.get("KRX_OPEN_API_KEY")
STK_BYDD_TRD_URL = "https://data-dbg.krx.co.kr/svc/apis/sto/stk_bydd_trd"
KSQ_BYDD_TRD_URL = "https://data-dbg.krx.co.kr/svc/apis/sto/ksq_bydd_trd"

SHEET_TO_PREFIX = {
    "개인_순매수": "indiv",
    "기관_순매수": "inst",
    "외국인_순매수": "foreign",
}


def find_manual_file():
    candidates = glob.glob(os.path.join(MANUAL_DIR, "*수급정리*.xls*"))
    candidates = [c for c in candidates if not os.path.basename(c).startswith("~$")]
    if not candidates:
        return None
    # 가장 최근에 수정된 파일 사용
    return max(candidates, key=os.path.getmtime)


def fetch_market_tickers(basDd):
    """해당 날짜 기준 코스피/코스닥 전종목 코드 집합. 최근 영업일이 없으면 며칠 앞으로 당겨본다."""
    if not KRX_OPEN_API_KEY:
        print("경고: KRX_OPEN_API_KEY가 없어 코스피/코스닥 분류를 할 수 없습니다.")
        return set(), set()

    headers = {"AUTH_KEY": KRX_OPEN_API_KEY}
    d = basDd
    for _ in range(10):
        kospi = requests.get(STK_BYDD_TRD_URL, params={"basDd": d}, headers=headers, timeout=20)
        kosdaq = requests.get(KSQ_BYDD_TRD_URL, params={"basDd": d}, headers=headers, timeout=20)
        kospi_rows = kospi.json().get("OutBlock_1", []) if kospi.status_code == 200 else []
        kosdaq_rows = kosdaq.json().get("OutBlock_1", []) if kosdaq.status_code == 200 else []
        if kospi_rows and kosdaq_rows:
            return {r["ISU_CD"] for r in kospi_rows}, {r["ISU_CD"] for r in kosdaq_rows}
        d = (datetime.strptime(d, "%Y%m%d") - timedelta(days=1)).strftime("%Y%m%d")
    return set(), set()


def build_market_by_name(ws_raw):
    """원본데이터 시트에서 종목명 -> KRX 6자리 코드 매핑을 만든다.

    read_only 워크시트에서는 .cell(row=, column=)로 임의 접근하면 매번 처음부터
    다시 스캔해서 사실상 멈춘 것처럼 느려진다 (수천 x 수천 셀에서 치명적).
    iter_rows()로 순차 스캔해야 실사용 가능한 속도가 나온다.
    """
    code_row, name_row = None, None
    for i, row in enumerate(ws_raw.iter_rows(min_row=1, max_row=9, values_only=True), start=1):
        if i == 8:
            code_row = row
        elif i == 9:
            name_row = row
    code_by_name = {}
    for c in range(1, len(name_row), 3):  # 0-index: 컬럼 B(=index1)부터 3칸 간격
        code = code_row[c] if c < len(code_row) else None
        name = name_row[c] if c < len(name_row) else None
        if name and code and name not in code_by_name:
            code_by_name[name] = str(code).lstrip("A")
    return code_by_name


def classify(code, kospi_set, kosdaq_set):
    if code in kospi_set:
        return "kospi"
    if code in kosdaq_set:
        return "kosdaq"
    return None


def sum_sheet(ws, code_by_name, kospi_set, kosdaq_set):
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    markets = []
    for name in header[1:]:
        code = code_by_name.get(name)
        markets.append(classify(code, kospi_set, kosdaq_set) if code else None)

    records = []
    for row in rows_iter:
        date_val = row[0]
        if date_val is None:
            continue
        kospi_sum = kosdaq_sum = total_sum = 0.0
        has_value = False
        for i, v in enumerate(row[1:]):
            if v is None:
                continue
            has_value = True
            total_sum += v
            m = markets[i] if i < len(markets) else None
            if m == "kospi":
                kospi_sum += v
            elif m == "kosdaq":
                kosdaq_sum += v
        if not has_value:
            continue
        records.append({"date": date_val, "kospi": kospi_sum, "kosdaq": kosdaq_sum, "total": total_sum})
    return pd.DataFrame(records)


def main():
    path = find_manual_file()
    if not path:
        print("data/manual/ 에 '수급정리' 엑셀 파일이 없습니다.")
        return

    print(f"파일 로딩 중: {os.path.basename(path)} (용량이 커서 시간이 좀 걸립니다)")
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)

    if "원본데이터" not in wb.sheetnames:
        print("경고: '원본데이터' 시트가 없어 코스피/코스닥 분류를 할 수 없습니다.")
        return
    code_by_name = build_market_by_name(wb["원본데이터"])
    print(f"종목-코드 매핑 {len(code_by_name)}개 확인")

    today_str = datetime.today().strftime("%Y%m%d")
    kospi_set, kosdaq_set = fetch_market_tickers(today_str)
    print(f"코스피 {len(kospi_set)}종목, 코스닥 {len(kosdaq_set)}종목 확인 (KRX Open API)")

    matched = sum(1 for n in code_by_name if classify(code_by_name[n], kospi_set, kosdaq_set))
    print(f"종목 분류 결과: {matched}/{len(code_by_name)}개 매칭됨")

    merged = None
    for sheet_name, prefix in SHEET_TO_PREFIX.items():
        if sheet_name not in wb.sheetnames:
            print(f"경고: 시트 '{sheet_name}' 없음, 스킵")
            continue
        print(f"집계 중: {sheet_name} ...")
        df = sum_sheet(wb[sheet_name], code_by_name, kospi_set, kosdaq_set)
        df = df.rename(columns={
            "kospi": f"{prefix}_net_kospi",
            "kosdaq": f"{prefix}_net_kosdaq",
            "total": f"{prefix}_net_total",
        })
        print(f"  {len(df)}행")
        merged = df if merged is None else merged.merge(df, on="date", how="outer")

    if merged is None or merged.empty:
        print("집계된 데이터가 없습니다.")
        return

    merged["date"] = pd.to_datetime(merged["date"])
    merged = merged.sort_values("date").reset_index(drop=True)

    if os.path.exists(OUT_PATH):
        existing = pd.read_csv(OUT_PATH, parse_dates=["date"])
        combined = pd.concat([existing, merged], ignore_index=True)
        combined = combined.drop_duplicates(subset="date", keep="last").sort_values("date").reset_index(drop=True)
    else:
        combined = merged

    os.makedirs(DATA_DIR, exist_ok=True)
    combined.to_csv(OUT_PATH, index=False, encoding="utf-8-sig")

    print(f"\n저장 완료: {OUT_PATH}")
    print(f"행 수: {len(combined)}, 기간: {combined['date'].min().date()} ~ {combined['date'].max().date()}")
    print(combined.tail(3).to_string(index=False))


if __name__ == "__main__":
    main()
